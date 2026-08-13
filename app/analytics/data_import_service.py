"""
管理员数据导入服务：从导出的 Excel 恢复数据

- 支持从导出目录（或用户指定路径）选择 xlsx 文件
- 用户可选择要导入的工作表（sheet）
- 用户可选择导入前需要清空的目标表
- 支持基础数据（9 表）与学生数据（背景资料/任务答案/反馈答案）

导入采用事务包裹：任一表导入失败则整体回滚。
"""

import os
from datetime import datetime

from openpyxl import load_workbook

from app.analytics.data_export_service import get_export_dir
from app.db import get_connection
from app.auth.auth_service import hash_password


# ================================================================
# 中文 -> 英文 反向映射（与导出映射对应）
# ================================================================

REV_ROLE = {'管理员': 'admin', '学生': 'student'}
REV_TASK_STATUS = {'草稿': 'draft', '已发布': 'published', '进行中': 'active', '已关闭': 'closed'}
REV_TASK_TYPE = {'调查任务': 'survey', '背景资料': 'background'}
REV_QUESTION_TYPE = {
    '单选题': 'single_choice',
    '多选题': 'multiple_choice',
    '开放式文本题': 'open',
}
REV_FEEDBACK_QUESTION_TYPE = {
    '单选题': 'radio',
    '多选题': 'checkbox',
    '开放式文本题': 'open',
}
REV_USER_TYPE = {'正式': 'formal', '试用': 'trial'}
REV_USER_STATUS = {'启用': 'active', '停用': 'disabled', '锁定': 'locked'}
YN_TO_INT = {'是': 1, '否': 0}


# ================================================================
# 工作表 schema 定义
# 每个条目：中文 sheet 名 -> {table, columns: [(中文表头, 数据库列, 转换器)]}
# ================================================================

SHEET_SCHEMAS = {
    '案例': {
        'table': 'cases',
        'columns': [
            ('案例ID', 'id', 'int'),
            ('案例标题', 'title', 'text'),
            ('案例正文', 'body', 'text'),
            ('主题', 'theme', 'text'),
            ('创建人账号', 'created_by', 'username'),
            ('创建时间', 'created_at', 'time'),
            ('更新时间', 'updated_at', 'time'),
        ],
    },
    '案例题目': {
        'table': 'case_questions',
        'columns': [
            ('题目ID', 'id', 'int'),
            ('案例ID', 'case_id', 'int'),
            ('题目文本', 'question_text', 'text'),
            ('题型', 'question_type', 'qtype'),
            ('选项(JSON)', 'options', 'text'),
            ('提示', 'hint', 'text'),
            ('排序号', 'sort_order', 'int'),
        ],
    },
    '任务': {
        'table': 'tasks',
        'columns': [
            ('任务ID', 'id', 'int'),
            ('任务名称', 'name', 'text'),
            ('任务描述', 'description', 'text'),
            ('开始时间', 'start_time', 'time'),
            ('结束时间', 'end_time', 'time'),
            ('状态', 'status', 'task_status'),
            ('任务类型', 'task_type', 'task_type'),
            ('创建时间', 'created_at', 'time'),
        ],
    },
    '任务关联案例': {
        'table': 'task_cases',
        'columns': [
            ('关联ID', 'id', 'int'),
            ('任务ID', 'task_id', 'int'),
            ('案例ID', 'case_id', 'int'),
            ('排序号', 'sort_order', 'int'),
        ],
    },
    '反馈任务': {
        'table': 'feedback_tasks',
        'columns': [
            ('反馈任务ID', 'id', 'int'),
            ('标题', 'title', 'text'),
            ('描述', 'description', 'text'),
            ('页面分类', 'page_category', 'text'),
            ('创建时间', 'created_at', 'time'),
        ],
    },
    '反馈题目': {
        'table': 'feedback_questions',
        'columns': [
            ('题目ID', 'id', 'int'),
            ('反馈任务ID', 'task_id', 'int'),
            ('题目文本', 'question_text', 'text'),
            ('题型', 'question_type', 'ftype'),
            ('排序号', 'sort_order', 'int'),
            ('是否必答', 'required', 'yn'),
        ],
    },
    '反馈题目选项': {
        'table': 'feedback_question_options',
        'columns': [
            ('选项ID', 'id', 'int'),
            ('题目ID', 'question_id', 'int'),
            ('选项文本', 'label', 'text'),
            ('选项值', 'value', 'int'),
            ('排序号', 'sort_order', 'int'),
            ('是否需评论', 'requires_comment', 'yn'),
            ('评论提示', 'comment_hint', 'text'),
        ],
    },
    '反馈任务映射': {
        'table': 'feedback_task_mappings',
        'columns': [
            ('映射ID', 'id', 'int'),
            ('反馈任务ID', 'task_id', 'int'),
            ('关联题目ID', 'survey_question_id', 'int'),
        ],
    },
    '账户信息': {
        'table': 'users',
        'columns': [
            ('用户ID', 'id', 'int'),
            ('用户名', 'username', 'text'),
            ('角色', 'role', 'role'),
            ('姓名', 'real_name', 'text'),
            ('班级', 'class_name', 'text'),
            ('学号', 'student_id', 'text'),
            ('用户类型', 'user_type', 'user_type'),
            ('状态', 'status', 'user_status'),
            ('需改密', 'must_change_password', 'yn'),
            ('创建时间', 'created_at', 'time'),
        ],
    },
    # ---- 学生数据 ----
    '背景资料': {
        'table': 'responses',
        'student': True,
    },
    '任务答案': {
        'table': 'responses',
        'student': True,
    },
    '反馈答案': {
        'table': 'feedback_responses',
        'student': True,
    },
}

# 清空顺序（子表在前，父表在后，避免外键约束冲突）
CLEAR_ORDER = [
    'response_details',
    'responses',
    'feedback_responses',
    'feedback_task_mappings',
    'feedback_question_options',
    'feedback_questions',
    'feedback_tasks',
    'task_cases',
    'case_questions',
    'tasks',
    'cases',
    'users',
]

# 导入顺序（父表在前，子表在后）
IMPORT_ORDER = [
    'users',
    'cases',
    'case_questions',
    'tasks',
    'task_cases',
    'feedback_tasks',
    'feedback_questions',
    'feedback_question_options',
    'feedback_task_mappings',
    'responses',
    'feedback_responses',
]

# 外键依赖：子表 -> 父表（用于清空时闭包扩展）
FK_DEPENDENCIES = {
    'case_questions': ['cases'],
    'task_cases': ['tasks', 'cases'],
    'cases': ['users'],                       # created_by
    'tasks': ['users'],                       # created_by
    'responses': ['tasks', 'cases', 'users'],
    'response_details': ['responses', 'case_questions'],
    'feedback_questions': ['feedback_tasks'],
    'feedback_question_options': ['feedback_questions'],
    'feedback_task_mappings': ['feedback_tasks', 'case_questions'],
    'feedback_responses': ['users', 'feedback_questions', 'feedback_question_options'],
}

# 反向外键：父表 -> 引用它的子表集合（用于清空闭包）
_PARENT_TO_CHILDREN = {}
for _child, _parents in FK_DEPENDENCIES.items():
    for _p in _parents:
        _PARENT_TO_CHILDREN.setdefault(_p, set()).add(_child)


def _expand_clear_tables(tables: list) -> list:
    """
    对用户选择清空的表做依赖闭包扩展：
    若清空父表，必须同时清空引用它的子表，否则外键约束报错
    返回按 CLEAR_ORDER 排序的表列表
    """
    result = set(tables)
    queue = list(tables)
    while queue:
        t = queue.pop()
        for child in _PARENT_TO_CHILDREN.get(t, ()):
            if child not in result:
                result.add(child)
                queue.append(child)
    return [t for t in CLEAR_ORDER if t in result]


# ================================================================
# 值转换
# ================================================================

def _to_int(value):
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _to_text(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_time(value):
    """时间转换为 SQLite 可存字符串；datetime 对象转文本"""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return str(value).strip()


def _map_value(value, mapping, default=None):
    """中文映射回英文；已知值返回映射结果，未知返回原值或默认"""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return default
    return mapping.get(s, s)


def _convert(value, converter):
    """按转换器类型处理单元格值"""
    if converter == 'int':
        return _to_int(value)
    if converter == 'text':
        return _to_text(value)
    if converter == 'time':
        return _to_time(value)
    if converter == 'yn':
        return YN_TO_INT.get(_to_text(value), 0)
    if converter == 'role':
        return _map_value(value, REV_ROLE, 'student')
    if converter == 'task_status':
        return _map_value(value, REV_TASK_STATUS, 'draft')
    if converter == 'task_type':
        return _map_value(value, REV_TASK_TYPE, 'survey')
    if converter == 'qtype':
        return _map_value(value, REV_QUESTION_TYPE, 'open')
    if converter == 'ftype':
        return _map_value(value, REV_FEEDBACK_QUESTION_TYPE, 'radio')
    if converter == 'user_type':
        return _map_value(value, REV_USER_TYPE, 'formal')
    if converter == 'user_status':
        return _map_value(value, REV_USER_STATUS, 'active')
    if converter == 'username':
        return _to_text(value)
    return value


# ================================================================
# 文件读取
# ================================================================

def list_excel_files(export_dir: str = None) -> list:
    """列出可导入的 xlsx 文件"""
    directory = export_dir or get_export_dir()
    if not os.path.isdir(directory):
        return []
    files = []
    for name in os.listdir(directory):
        if name.lower().endswith('.xlsx') and not name.startswith('~$'):
            files.append(name)
    return sorted(files)


def read_workbook_sheets(filepath: str) -> dict:
    """
    读取工作簿中可导入的工作表信息
    返回 {sheet名: {'table': 目标表, 'rows': 数据行数, 'headers': [中文表头], 'student': bool}}
    仅返回导出服务定义过的工作表，其余忽略
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        result = {}
        for ws in wb.worksheets:
            schema = SHEET_SCHEMAS.get(ws.title)
            if not schema:
                continue
            headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            data_rows = max(ws.max_row - 1, 0)
            result[ws.title] = {
                'table': schema['table'],
                'rows': data_rows,
                'headers': [str(h) if h is not None else '' for h in headers],
                'student': bool(schema.get('student')),
            }
        return result
    finally:
        wb.close()


def _iter_rows(ws):
    """跳过表头，迭代数据行（列表形式）"""
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if any(v is not None and str(v).strip() != '' for v in row):
            yield i, list(row)


# ================================================================
# 清空表
# ================================================================

def clear_tables(table_names: list, commit: bool = True) -> dict:
    """
    按依赖顺序清空指定表（只清勾选的表，父表被清时子表先清）
    返回 {表名: 删除行数}
    commit=False 时由调用方统一提交（用于整体事务）
    """
    if not table_names:
        return {}
    with get_connection() as conn:
        cursor = conn.cursor()
        stats = {}
        # 依赖闭包扩展 + 按 CLEAR_ORDER 排序（子表先删，父表后删）
        ordered = _expand_clear_tables(table_names)
        for table in ordered:
            cursor.execute(f"DELETE FROM {table}")
            stats[table] = cursor.rowcount
        if commit:
            conn.commit()
        return stats


# ================================================================
# 通用表导入
# ================================================================

def _id_exists(cursor, table, pk_id):
    cursor.execute(f"SELECT 1 FROM {table} WHERE id = :pk", {'pk': pk_id})
    return cursor.fetchone() is not None


def _username_to_id(cursor, username):
    """用户名 -> users.id，找不到返回 None"""
    name = _to_text(username)
    if not name:
        return None
    cursor.execute("SELECT id FROM users WHERE username = :u", {'u': name})
    row = cursor.fetchone()
    return row[0] if row else None


def import_sheet(cursor, ws, sheet_name: str) -> dict:
    """
    导入单个工作表到目标表
    返回 {'imported': n, 'skipped': m, 'table': 表名}
    已存在同 ID 记录则跳过（防止重复导入报错）
    """
    schema = SHEET_SCHEMAS[sheet_name]
    table = schema['table']
    columns = schema['columns']

    # 建立 中文表头 -> 数据库列 的索引
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c) if c is not None else '' for c in header_row]
    col_map = {}
    for (header, db_col, converter) in columns:
        if header in headers:
            col_map[header] = (db_col, converter)

    col_names = [db for (_, db, _) in columns]
    # 账户信息缺密码列，导入时追加（生成默认密码）
    if table == 'users':
        for extra in ('password_hash', 'salt'):
            if extra not in col_names:
                col_names.append(extra)
    insert_sql = (
        f"INSERT INTO {table} ({', '.join(col_names)}) "
        f"VALUES ({', '.join(':' + c for c in col_names)})"
    )

    imported = 0
    skipped = 0
    for _, row in _iter_rows(ws):
        record = {}
        for i, header in enumerate(headers):
            if header in col_map and i < len(row):
                db_col, converter = col_map[header]
                record[db_col] = _convert(row[i], converter)

        # 特殊处理：账户信息缺密码，生成默认密码（123456，强制改密）
        if table == 'users':
            record['password_hash'], record['salt'] = hash_password('123456')
            record['must_change_password'] = 1

        # 特殊处理：cases.created_by 导出为用户名，解析为 users.id
        if table == 'cases' and record.get('created_by'):
            record['created_by'] = _username_to_id(cursor, record['created_by'])

        pk = record.get('id')
        if pk is not None and _id_exists(cursor, table, pk):
            skipped += 1
            continue
        cursor.execute(insert_sql, record)
        imported += 1

    return {'imported': imported, 'skipped': skipped, 'table': table}


# ================================================================
# 学生数据导入（responses / feedback_responses）
# ================================================================

def _read_headers(ws):
    """读取工作表表头（字符串列表），供按列名定位"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(c) if c is not None else '' for c in header_row]


def _col_index(headers, name):
    """按表头名取列索引，找不到返回 None"""
    try:
        return headers.index(name)
    except ValueError:
        return None


def import_student_sheet(cursor, ws, sheet_name: str) -> dict:
    """
    导入学生数据工作表：
    - 背景资料 / 任务答案 -> responses + response_details（按 用户名+任务+案例 分组）
    - 反馈答案 -> feedback_responses
    """
    sheet_type = {'背景资料': 'background', '任务答案': 'answer', '反馈答案': 'feedback'}[sheet_name]
    headers = _read_headers(ws)
    idx = {
        'username': _col_index(headers, '用户名'),
        'task_name': _col_index(headers, '任务名称'),
        'case_title': _col_index(headers, '案例标题'),
        'question_text': _col_index(headers, '题目文本'),
        'answer': _col_index(headers, '答案'),
        'submitted_at': _col_index(headers, '提交时间'),
        'task_title': _col_index(headers, '反馈任务'),
        'option_label': _col_index(headers, '所选选项'),
        'comment_text': _col_index(headers, '评论文本'),
        'created_at': _col_index(headers, '提交时间'),
    }
    imported = 0
    skipped = 0

    if sheet_type in ('background', 'answer'):
        groups = {}  # (student_id, task_id, case_id) -> response_id
        for _, row in _iter_rows(ws):
            def get(col):
                i = idx.get(col)
                if i is None or i >= len(row):
                    return None
                return row[i]

            username = _to_text(get('username'))
            task_name = _to_text(get('task_name'))
            case_title = _to_text(get('case_title'))
            question_text = _to_text(get('question_text'))
            raw_answer = get('answer')
            answer = str(raw_answer) if raw_answer is not None else ''
            submitted_at = _to_time(get('submitted_at'))
            if not (username and task_name and case_title and question_text):
                skipped += 1
                continue

            student_id = _username_to_id(cursor, username)
            if student_id is None:
                skipped += 1
                continue

            # 任务
            cursor.execute(
                "SELECT id FROM tasks WHERE name = :n", {'n': task_name})
            task_row = cursor.fetchone()
            # 案例
            cursor.execute(
                "SELECT id FROM cases WHERE title = :t", {'t': case_title})
            case_row = cursor.fetchone()
            if task_row is None or case_row is None:
                skipped += 1
                continue
            task_id, case_id = task_row[0], case_row[0]

            # 题目（需结合案例）
            cursor.execute(
                "SELECT id FROM case_questions WHERE case_id = :cid AND question_text = :qt",
                {'cid': case_id, 'qt': question_text})
            q_row = cursor.fetchone()
            if q_row is None:
                skipped += 1
                continue
            question_id = q_row[0]

            key = (student_id, task_id, case_id)
            if key not in groups:
                # 创建 response
                cursor.execute("""
                    SELECT id FROM responses WHERE task_id = :t AND case_id = :c AND student_id = :s
                """, {'t': task_id, 'c': case_id, 's': student_id})
                existing = cursor.fetchone()
                if existing:
                    groups[key] = existing[0]
                else:
                    cursor.execute("""
                        INSERT INTO responses (task_id, case_id, student_id, status, submitted_at, updated_at)
                        VALUES (:t, :c, :s, 'submitted', :sa, CURRENT_TIMESTAMP)
                    """, {'t': task_id, 'c': case_id, 's': student_id, 'sa': submitted_at})
                    groups[key] = cursor.lastrowid

            response_id = groups[key]
            # 题目答案去重（同一题重复行跳过）
            cursor.execute("""
                SELECT 1 FROM response_details WHERE response_id = :r AND question_id = :q
            """, {'r': response_id, 'q': question_id})
            if cursor.fetchone():
                skipped += 1
                continue
            cursor.execute("""
                INSERT INTO response_details (response_id, question_id, answer) VALUES (:r, :q, :a)
            """, {'r': response_id, 'q': question_id, 'a': answer})
            imported += 1
    else:
        # 反馈答案 -> feedback_responses
        for _, row in _iter_rows(ws):
            def get(col):
                i = idx.get(col)
                if i is None or i >= len(row):
                    return None
                return row[i]

            username = _to_text(get('username'))
            task_title = _to_text(get('task_title'))
            question_text = _to_text(get('question_text'))
            option_label = _to_text(get('option_label'))
            comment_text = _to_text(get('comment_text'))
            created_at = _to_time(get('created_at'))
            if not (username and task_title and question_text):
                skipped += 1
                continue

            student_id = _username_to_id(cursor, username)
            if student_id is None:
                skipped += 1
                continue

            cursor.execute(
                "SELECT id FROM feedback_tasks WHERE title = :t", {'t': task_title})
            task_row = cursor.fetchone()
            if task_row is None:
                skipped += 1
                continue
            task_id = task_row[0]

            cursor.execute(
                "SELECT id FROM feedback_questions WHERE task_id = :t AND question_text = :qt",
                {'t': task_id, 'qt': question_text})
            q_row = cursor.fetchone()
            if q_row is None:
                skipped += 1
                continue
            question_id = q_row[0]

            option_id = None
            if option_label:
                cursor.execute(
                    "SELECT id FROM feedback_question_options WHERE question_id = :q AND label = :l",
                    {'q': question_id, 'l': option_label})
                opt_row = cursor.fetchone()
                option_id = opt_row[0] if opt_row else None

            cursor.execute("""
                INSERT INTO feedback_responses
                    (student_id, survey_question_id, feedback_question_id, selected_option_id, comment_text, created_at)
                VALUES (:s, NULL, :q, :o, :c, :ca)
            """, {'s': student_id, 'q': question_id, 'o': option_id, 'c': comment_text, 'ca': created_at})
            imported += 1

    return {'imported': imported, 'skipped': skipped, 'table': 'responses' if sheet_type != 'feedback' else 'feedback_responses'}


# ================================================================
# 对外接口
# ================================================================

def do_import(filepath: str, sheets_to_import: list, tables_to_clear: list) -> dict:
    """
    执行导入：
    1. 清空选中的目标表
    2. 导入选中的工作表
    整体事务，失败回滚
    返回 {'cleared': {表: 行数}, 'imported': {sheet: {imported, skipped}}, 'error': str|None}
    """
    if not os.path.exists(filepath):
        return {'cleared': {}, 'imported': {}, 'error': f'文件不存在: {filepath}'}
    if not sheets_to_import:
        return {'cleared': {}, 'imported': {}, 'error': '请选择要导入的工作表'}

    wb = load_workbook(filepath, data_only=True)
    cleared = {}
    imported = {}
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            # 1. 清空表（不提交，与导入同事务，失败整体回滚）
            cleared = clear_tables(tables_to_clear, commit=False)
            # 2. 按依赖顺序导入（父表先导入，保证外键成立）
            def order_key(name):
                table = SHEET_SCHEMAS[name]['table']
                return IMPORT_ORDER.index(table) if table in IMPORT_ORDER else len(IMPORT_ORDER)

            for sheet_name in sorted(sheets_to_import, key=order_key):
                if sheet_name not in SHEET_SCHEMAS or sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                schema = SHEET_SCHEMAS[sheet_name]
                if schema.get('student'):
                    imported[sheet_name] = import_student_sheet(cursor, ws, sheet_name)
                else:
                    imported[sheet_name] = import_sheet(cursor, ws, sheet_name)
            conn.commit()
            return {'cleared': cleared, 'imported': imported, 'error': None}
    except Exception as ex:
        try:
            conn.rollback()
        except Exception:
            pass
        return {'cleared': {}, 'imported': {}, 'error': str(ex)}
    finally:
        wb.close()
