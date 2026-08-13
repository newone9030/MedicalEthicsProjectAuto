"""
管理员数据导出服务：基础数据 / 学生数据 Excel 导出

- 基础数据：案例、案例题目、任务、任务关联案例、反馈任务、反馈题目、反馈题目选项、
            反馈任务映射、账户信息（排除密码哈希/盐）
- 学生数据：背景资料答案、任务答案、反馈答案
- 导出目录可由管理员设置，持久化保存于 export_config.json

所有工作表表头均使用中文。
"""

import json
import os
from datetime import datetime

from app.config import BASE_DIR
from app.db import get_connection

# 导出目录配置文件（持久化用户设置的目录）
EXPORT_CONFIG_PATH = os.path.join(BASE_DIR, 'export_config.json')
DEFAULT_EXPORT_DIR = os.environ.get(
    'EXPORT_DIR',
    os.path.join(BASE_DIR, 'exports')
)

# 题目类型映射
QUESTION_TYPE_CN = {
    'single_choice': '单选题',
    'multiple_choice': '多选题',
    'open': '开放式文本题',
}

# 反馈题目题型映射
FEEDBACK_QUESTION_TYPE_CN = {
    'radio': '单选题',
    'checkbox': '多选题',
    'open': '开放式文本题',
}

# 角色 / 状态 / 用户类型 / 任务类型中文映射
ROLE_CN = {'admin': '管理员', 'student': '学生'}
STATUS_CN = {'draft': '草稿', 'published': '已发布', 'active': '进行中', 'closed': '已关闭'}
USER_TYPE_CN = {'formal': '正式', 'trial': '试用'}
USER_STATUS_CN = {'active': '启用', 'disabled': '停用', 'locked': '锁定'}
TASK_TYPE_CN = {'survey': '调查任务', 'background': '背景资料'}


# ================================================================
# 导出目录配置
# ================================================================

def get_export_dir() -> str:
    """获取导出目录（读取持久化配置，不存在则返回默认目录）"""
    try:
        if os.path.exists(EXPORT_CONFIG_PATH):
            with open(EXPORT_CONFIG_PATH, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            path = cfg.get('export_dir')
            if path and isinstance(path, str):
                return os.path.normpath(path)
    except (json.JSONDecodeError, OSError, TypeError):
        pass
    return DEFAULT_EXPORT_DIR


def set_export_dir(path: str) -> str:
    """设置并持久化导出目录，返回规范化后的路径"""
    path = (path or '').strip().strip('"').strip("'")
    if not path:
        raise ValueError('导出目录不能为空')
    path = os.path.normpath(path)
    try:
        os.makedirs(path, exist_ok=True)
    except OSError as ex:
        raise ValueError(f'无法创建目录: {path}（{ex}）')
    with open(EXPORT_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump({'export_dir': path}, f, ensure_ascii=False, indent=2)
    return path


def ensure_export_dir() -> str:
    """确保导出目录存在，返回路径"""
    path = get_export_dir()
    os.makedirs(path, exist_ok=True)
    return path


# ================================================================
# 通用工具
# ================================================================

def _fmt_time(value):
    """时间格式化，避免 None / datetime 写入问题"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return str(value)


def _cn(value, mapping):
    """按映射表转为中文，未知值原样返回"""
    return mapping.get(value, value)


def format_answer(raw) -> str:
    """
    将答案转为可读文本（仅供预览，导出时保留原始格式以便无损导入）：
    - 普通任务单选/开放题：直接返回文本
    - 多选：JSON 数组 -> 顿号分隔
    - 背景资料（含说明）：JSON 对象 -> "答案；说明：xxx"
    """
    if raw is None:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    try:
        data = json.loads(s)
    except (json.JSONDecodeError, TypeError):
        return s
    if isinstance(data, dict):
        parts = []
        answer = data.get('answer')
        explanation = data.get('explanation')
        if answer is None:
            # 任务作答端的结构化答案：{'option': ..., 'open_text': ...} 或 {'options': [...], 'open_text': ...}
            option = data.get('option')
            options = data.get('options')
            open_text = data.get('open_text')
            if option:
                parts.append(str(option))
            if options:
                parts.append('、'.join(str(x) for x in options) if isinstance(options, list) else str(options))
            if open_text:
                parts.append(f'补充：{open_text}')
        else:
            if answer:
                parts.append(str(answer))
            if explanation:
                parts.append(f'说明：{explanation}')
        return '；'.join(parts) if parts else s
    if isinstance(data, list):
        return '、'.join(str(x) for x in data)
    return str(data)


def _raw_answer(raw) -> str:
    """导出时保留答案原始存储格式，保证可无损导入"""
    return '' if raw is None else str(raw)


def _auto_width(ws, max_width: int = 50):
    """根据内容自动调整列宽（粗略估算）"""
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                # 中文字符按 2 个宽度估算
                length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), max_width)


def _write_sheet(ws, headers: list, rows: list, freeze: str = 'A2'):
    """写入表头与数据，设置表头样式并冻结首行"""
    from openpyxl.styles import Font, PatternFill

    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell.alignment.copy(wrap_text=True, vertical='center')
    ws.freeze_panes = freeze


# ================================================================
# 基础数据导出
# ================================================================

def export_basic_data(filepath: str) -> dict:
    """导出基础数据 Excel，返回 {sheet: 行数} 统计"""
    from openpyxl import Workbook

    wb = Workbook()
    first_sheet = True
    stats = {}

    with get_connection() as conn:
        cursor = conn.cursor()

        # ---- 1. 案例 ----
        cursor.execute("""
            SELECT c.id, c.title, c.body, c.theme, u.real_name, u.username, c.created_at, c.updated_at
            FROM cases c
            LEFT JOIN users u ON c.created_by = u.id
            ORDER BY c.id
        """)
        rows = [
            [r[0], r[1], r[2] or '', r[3] or '', r[4] or r[5] or '', _fmt_time(r[6]), _fmt_time(r[7])]
            for r in cursor.fetchall()
        ]
        ws = wb.active if first_sheet else wb.create_sheet()
        first_sheet = False
        ws.title = '案例'
        _write_sheet(ws, ['案例ID', '案例标题', '案例正文', '主题', '创建人', '创建人账号', '创建时间', '更新时间'], rows)
        _auto_width(ws)
        stats['案例'] = len(rows)

        # ---- 2. 案例题目 ----
        cursor.execute("""
            SELECT q.id, q.case_id, c.title, q.question_text, q.question_type, q.options, q.hint, q.sort_order
            FROM case_questions q
            JOIN cases c ON q.case_id = c.id
            ORDER BY q.case_id, q.sort_order
        """)
        rows = [
            [r[0], r[1], r[2], r[3], _cn(r[4], QUESTION_TYPE_CN), r[5] or '', r[6] or '', r[7]]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '案例题目'
        _write_sheet(ws, ['题目ID', '案例ID', '所属案例', '题目文本', '题型', '选项(JSON)', '提示', '排序号'], rows)
        _auto_width(ws)
        stats['案例题目'] = len(rows)

        # ---- 3. 任务 ----
        cursor.execute("""
            SELECT t.id, t.name, t.description, t.start_time, t.end_time, t.status, t.task_type,
                   u.real_name, u.username, t.created_at
            FROM tasks t
            LEFT JOIN users u ON t.created_by = u.id
            ORDER BY t.id
        """)
        rows = [
            [r[0], r[1], r[2] or '', _fmt_time(r[3]), _fmt_time(r[4]),
             _cn(r[5], STATUS_CN), _cn(r[6], TASK_TYPE_CN), r[7] or r[8] or '', _fmt_time(r[9])]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '任务'
        _write_sheet(ws, ['任务ID', '任务名称', '任务描述', '开始时间', '结束时间', '状态', '任务类型', '创建人', '创建时间'], rows)
        _auto_width(ws)
        stats['任务'] = len(rows)

        # ---- 4. 任务关联案例 ----
        cursor.execute("""
            SELECT tc.id, tc.task_id, t.name, tc.case_id, c.title, tc.sort_order
            FROM task_cases tc
            JOIN tasks t ON tc.task_id = t.id
            JOIN cases c ON tc.case_id = c.id
            ORDER BY tc.task_id, tc.sort_order
        """)
        rows = [
            [r[0], r[1], r[2], r[3], r[4], r[5]]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '任务关联案例'
        _write_sheet(ws, ['关联ID', '任务ID', '任务名称', '案例ID', '案例标题', '排序号'], rows)
        _auto_width(ws)
        stats['任务关联案例'] = len(rows)

        # ---- 5. 反馈任务 ----
        cursor.execute("""
            SELECT id, title, description, page_category, created_at
            FROM feedback_tasks
            ORDER BY id
        """)
        rows = [
            [r[0], r[1], r[2] or '', r[3] or '', _fmt_time(r[4])]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '反馈任务'
        _write_sheet(ws, ['反馈任务ID', '标题', '描述', '页面分类', '创建时间'], rows)
        _auto_width(ws)
        stats['反馈任务'] = len(rows)

        # ---- 6. 反馈题目 ----
        cursor.execute("""
            SELECT q.id, q.task_id, t.title, q.question_text, q.question_type, q.sort_order, q.required
            FROM feedback_questions q
            JOIN feedback_tasks t ON q.task_id = t.id
            ORDER BY q.task_id, q.sort_order
        """)
        rows = [
            [r[0], r[1], r[2], r[3], _cn(r[4], FEEDBACK_QUESTION_TYPE_CN), r[5], '是' if r[6] else '否']
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '反馈题目'
        _write_sheet(ws, ['题目ID', '反馈任务ID', '所属反馈任务', '题目文本', '题型', '排序号', '是否必答'], rows)
        _auto_width(ws)
        stats['反馈题目'] = len(rows)

        # ---- 7. 反馈题目选项 ----
        cursor.execute("""
            SELECT o.id, o.question_id, q.question_text, o.label, o.value, o.sort_order,
                   o.requires_comment, o.comment_hint
            FROM feedback_question_options o
            JOIN feedback_questions q ON o.question_id = q.id
            ORDER BY o.question_id, o.sort_order
        """)
        rows = [
            [r[0], r[1], r[2], r[3], r[4], r[5], '是' if r[6] else '否', r[7] or '']
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '反馈题目选项'
        _write_sheet(ws, ['选项ID', '题目ID', '所属题目', '选项文本', '选项值', '排序号', '是否需评论', '评论提示'], rows)
        _auto_width(ws)
        stats['反馈题目选项'] = len(rows)

        # ---- 8. 反馈任务映射 ----
        cursor.execute("""
            SELECT m.id, m.task_id, t.title, m.survey_question_id, q.question_text
            FROM feedback_task_mappings m
            JOIN feedback_tasks t ON m.task_id = t.id
            LEFT JOIN case_questions q ON m.survey_question_id = q.id
            ORDER BY m.task_id, m.survey_question_id
        """)
        rows = [
            [r[0], r[1], r[2], r[3], r[4] or '']
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '反馈任务映射'
        _write_sheet(ws, ['映射ID', '反馈任务ID', '反馈任务标题', '关联题目ID', '关联题目文本'], rows)
        _auto_width(ws)
        stats['反馈任务映射'] = len(rows)

        # ---- 9. 账户信息（排除 password_hash / salt）----
        cursor.execute("""
            SELECT id, username, role, real_name, class_name, student_id,
                   user_type, status, must_change_password, created_at
            FROM users
            ORDER BY id
        """)
        rows = [
            [r[0], r[1], _cn(r[2], ROLE_CN), r[3] or '', r[4] or '', r[5] or '',
             _cn(r[6], USER_TYPE_CN), _cn(r[7], USER_STATUS_CN), '是' if r[8] else '否', _fmt_time(r[9])]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '账户信息'
        _write_sheet(ws, ['用户ID', '用户名', '角色', '姓名', '班级', '学号', '用户类型', '状态', '需改密', '创建时间'], rows)
        _auto_width(ws)
        stats['账户信息'] = len(rows)

    wb.save(filepath)
    return stats


# ================================================================
# 学生数据导出
# ================================================================

def export_student_data(filepath: str) -> dict:
    """导出学生数据 Excel，返回 {sheet: 行数} 统计"""
    from openpyxl import Workbook

    wb = Workbook()
    first_sheet = True
    stats = {}

    with get_connection() as conn:
        cursor = conn.cursor()

        # ---- 1. 背景资料答案 ----
        cursor.execute("""
            SELECT u.username, u.real_name, u.class_name, u.student_id,
                   t.name, c.title, q.question_text, d.answer, r.submitted_at
            FROM responses r
            JOIN users u ON r.student_id = u.id
            JOIN tasks t ON r.task_id = t.id
            JOIN task_cases tc ON r.task_id = tc.task_id AND r.case_id = tc.case_id
            JOIN cases c ON r.case_id = c.id
            JOIN response_details d ON d.response_id = r.id
            JOIN case_questions q ON d.question_id = q.id
            WHERE r.status = 'submitted' AND t.task_type = 'background'
            ORDER BY u.id, t.id, c.id, q.sort_order
        """)
        rows = [
            [r[0], r[1] or '', r[2] or '', r[3] or '', r[4], r[5], r[6], _raw_answer(r[7]), _fmt_time(r[8])]
            for r in cursor.fetchall()
        ]
        ws = wb.active if first_sheet else wb.create_sheet()
        first_sheet = False
        ws.title = '背景资料'
        _write_sheet(ws, ['用户名', '姓名', '班级', '学号', '任务名称', '案例标题', '题目文本', '答案', '提交时间'], rows)
        _auto_width(ws)
        stats['背景资料'] = len(rows)

        # ---- 2. 任务答案 ----
        cursor.execute("""
            SELECT u.username, u.real_name, u.class_name, u.student_id,
                   t.name, c.title, q.question_text, q.question_type, d.answer, r.submitted_at
            FROM responses r
            JOIN users u ON r.student_id = u.id
            JOIN tasks t ON r.task_id = t.id
            JOIN task_cases tc ON r.task_id = tc.task_id AND r.case_id = tc.case_id
            JOIN cases c ON r.case_id = c.id
            JOIN response_details d ON d.response_id = r.id
            JOIN case_questions q ON d.question_id = q.id
            WHERE r.status = 'submitted' AND t.task_type != 'background'
            ORDER BY u.id, t.id, c.id, q.sort_order
        """)
        rows = [
            [r[0], r[1] or '', r[2] or '', r[3] or '', r[4], r[5], r[6],
             _cn(r[7], QUESTION_TYPE_CN), _raw_answer(r[8]), _fmt_time(r[9])]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '任务答案'
        _write_sheet(ws, ['用户名', '姓名', '班级', '学号', '任务名称', '案例标题', '题目文本', '题型', '答案', '提交时间'], rows)
        _auto_width(ws)
        stats['任务答案'] = len(rows)

        # ---- 3. 反馈答案 ----
        cursor.execute("""
            SELECT u.username, u.real_name, u.class_name, u.student_id,
                   t.title, q.question_text, o.label, fr.comment_text, fr.created_at
            FROM feedback_responses fr
            JOIN users u ON fr.student_id = u.id
            JOIN feedback_questions q ON fr.feedback_question_id = q.id
            JOIN feedback_tasks t ON q.task_id = t.id
            LEFT JOIN feedback_question_options o ON fr.selected_option_id = o.id
            ORDER BY u.id, t.id, q.sort_order, fr.created_at
        """)
        rows = [
            [r[0], r[1] or '', r[2] or '', r[3] or '', r[4], r[5],
             r[6] or '', r[7] or '', _fmt_time(r[8])]
            for r in cursor.fetchall()
        ]
        ws = wb.create_sheet()
        ws.title = '反馈答案'
        _write_sheet(ws, ['用户名', '姓名', '班级', '学号', '反馈任务', '题目文本', '所选选项', '评论文本', '提交时间'], rows)
        _auto_width(ws)
        stats['反馈答案'] = len(rows)

    wb.save(filepath)
    return stats


# ================================================================
# 对外封装
# ================================================================

def do_export(scope: str, export_dir: str = None) -> dict:
    """
    执行导出，返回 {filepath, filename, stats}
    scope: 'basic' 基础数据 / 'student' 学生数据 / 'all' 全部
    """
    target_dir = export_dir or ensure_export_dir()
    os.makedirs(target_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results = []

    if scope in ('basic', 'all'):
        filename = f'基础数据_{timestamp}.xlsx'
        filepath = os.path.join(target_dir, filename)
        stats = export_basic_data(filepath)
        results.append({'scope': 'basic', 'filename': filename, 'filepath': filepath, 'stats': stats})

    if scope in ('student', 'all'):
        filename = f'学生数据_{timestamp}.xlsx'
        filepath = os.path.join(target_dir, filename)
        stats = export_student_data(filepath)
        results.append({'scope': 'student', 'filename': filename, 'filepath': filepath, 'stats': stats})

    return results
